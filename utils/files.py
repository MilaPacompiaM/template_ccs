import csv

from django.template.loader import get_template

from openpyxl import load_workbook, Workbook
from wkhtmltopdf.utils import render_pdf_from_template


class ExcelManager:

    @staticmethod
    def read_values(excel_file, datamap):
        """
        Reads cells specified in `datamap` from `excel_file`

        :param excel_file: binary Excel file
        :param datamap: list of fields to be read

        :return: dict
        """

        file = load_workbook(excel_file)
        valueset = {}
        for field in datamap:
            row, col = field.get('position')
            sheet = file.get_sheet_by_name(field.get('sheet'))
            value = sheet.cell(row, col).value
            valueset[field.get('name')] = value
        return valueset

    @staticmethod
    def write(values, datamap):
        """
        Writes static and dynamic data structured in `datamap`

        :param values: dict with dynamic or variable data
        :param datamap: dict with structure of data

        :return: Workbook
        """
        wb = Workbook()
        for field in datamap:
            key = field.get('key', None)
            value = values.get(key, None) if key else field.get('value', None)
            row, col = field.get('position')
            main_sheet = wb.worksheets[0]
            main_sheet.cell(row=row, column=col).value = value
        return wb


class CsvManager:
    EMPTY_VALUE = ""
    MAX_ROWS = 10
    MAX_COLS = 10

    @staticmethod
    def read_values(csv_file, datamap):
        """
        Reads cells specified in `datamap` from `csv_file`

        :param csv_file: binary Excel file
        :param datamap: list of fields to be read

        :return: dict
        """

        file = csv_file.read().decode('UTF-8')
        rows = file.split("\n")
        data = [row.split(",") for row in rows]
        valueset = {}
        for field in datamap:
            row, col = field.get('position')
            # sheet name to filter are not available for csv
            if len(data) < row or len(data[row - 1]) < col:
                continue
            value = data[row - 1][col - 1]
            valueset[field.get('name')] = value
        return valueset

    @staticmethod
    def write(values, datamap):
        """
        Writes static and dynamic data structured in `datamap`

        :param values: dict with dynamic or variable data
        :param datamap: dict with structure of data

        :return: list of list (rows with column values)
        """

        data = [
            [CsvManager.EMPTY_VALUE for _ in range(CsvManager.MAX_COLS)]
            for _ in range(CsvManager.MAX_ROWS)
        ]
        for field in datamap:
            row, col = field.get('position')
            key = field.get('key', None)
            value = values.get(key, None) if key else field.get('value', None)
            if row > len(data) or col > len(data[0]):
                continue
            data[row - 1][col - 1] = value
        return data


class PDFManager:
    @staticmethod
    def write(values, template_name='demo.html', request=None):
        """
        Writes static and dynamic data structured in `datamap`

        :param values: dict with dynamic or variable data
        :param datamap: dict with structure of data

        :return: pdf object
        """

        # html = render_to_string(template, context=values)
        template = get_template(template_name)
        return render_pdf_from_template(
            input_template=template,
            header_template=None,
            footer_template=None,
            context=values,
            request=request
        )
